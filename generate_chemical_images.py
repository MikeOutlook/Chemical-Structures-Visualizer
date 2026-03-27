import pandas as pd
from rdkit import Chem
from rdkit.Chem import Draw
import openpyxl
from openpyxl.drawing.image import Image
import os

# Read CSV file
csv_path = 'chemical_structures_data.csv'
df = pd.read_csv(csv_path)

# Create output directory for images
img_dir = 'chemical_images'
os.makedirs(img_dir, exist_ok=True)

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Chemical Structures'

# Write headers
ws['A1'] = 'Index'
ws['B1'] = 'SMILES'
ws['C1'] = 'Structure'

# Generate images for each SMILES
for idx, row in df.iterrows():
    index = row['Index']
    smiles = row['SMILES']

    # Generate molecule from SMILES
    mol = Chem.MolFromSmiles(smiles)

    if mol is not None:
        # Generate image
        img_filename = 'compound_' + str(index) + '.png'
        img_path = os.path.join(img_dir, img_filename)
        img = Draw.MolToImage(mol, size=(300, 200))
        img.save(img_path)

        # Write data to Excel
        row_num = idx + 2
        ws.cell(row=row_num, column=1, value=index)
        ws.cell(row=row_num, column=2, value=smiles)

        # Add image to column C (right of SMILES)
        img_obj = Image(img_path)
        img_obj.width = 150
        img_obj.height = 100
        ws.add_image(img_obj, f'C{row_num}')

        print(f"Processed compound {index}: {smiles[:30]}...")
    else:
        print(f"Failed to parse SMILES for index {index}: {smiles[:30]}...")

# Adjust column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 20

# Save Excel file
output_path = 'chemical_structures_with_images.xlsx'
wb.save(output_path)
print(f"\nExcel file saved to: {output_path}")
print(f"Images saved to: {img_dir}")
